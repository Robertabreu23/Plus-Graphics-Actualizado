from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import sqlite3
from datetime import datetime, timedelta
from models import init_db
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Credenciales SOLO desde variables de entorno - SIN valores por defecto expuestos
ADMIN_EMAIL = os.getenv('ADMIN_EMAIL')
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD')
EMPLOYEE_EMAIL = os.getenv('EMPLOYEE_EMAIL')
EMPLOYEE_PASSWORD = os.getenv('EMPLOYEE_PASSWORD')

# Verificacion de seguridad al iniciar aplicacion
def verificar_variables_entorno():
    """Verificar que todas las variables criticas esten configuradas"""
    variables_requeridas = [
        'ADMIN_EMAIL', 'ADMIN_PASSWORD', 
        'EMPLOYEE_EMAIL', 'EMPLOYEE_PASSWORD'
    ]
    
    faltantes = [var for var in variables_requeridas if not os.getenv(var)]
    
    if faltantes:
        print(f"ERROR: Variables de entorno faltantes: {faltantes}")
        print("Sistema no puede iniciar sin credenciales configuradas")
        return False
    
    print("✅ Variables de entorno de seguridad configuradas correctamente")
    return True

def check_login_credentials(email, password):
    """Verificar credenciales desde variables de entorno seguras"""
    if not verificar_variables_entorno():
        raise Exception("Variables de entorno no configuradas")
    
    valid_credentials = {
        ADMIN_EMAIL: ADMIN_PASSWORD,
        EMPLOYEE_EMAIL: EMPLOYEE_PASSWORD
    }
    
    return email in valid_credentials and valid_credentials.get(email) == password

# Conexión a la base de datos
def get_db_connection():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row  # Para obtener diccionarios en lugar de tuplas
    return conn

app = Flask(__name__)

# CORS configurado para Vercel + desarrollo local
CORS(app, origins=[
    "http://localhost:3000",           # Desarrollo local
    "https://*.vercel.app",            # Cualquier subdominio Vercel
    "https://vercel.app",
    "https://*.onrender.com",          # Render por si acaso
    "https://plus-graphics.onrender.com"  # URL específica producción
])

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response
# -------------------- RUTAS DE AUTENTICACIÓN --------------------
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    
    # Verificar credenciales usando variables de entorno
    if check_login_credentials(data['email'], data['password']):
        # Determinar rol basado en email
        role = 'admin' if data['email'] == ADMIN_EMAIL else 'employee'
        name = 'Administrador' if data['email'] == ADMIN_EMAIL else 'Empleado'
        
        user_data = {
            'id': 1 if role == 'admin' else 2,
            'name': name,
            'email': data['email'],
            'role': role,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return jsonify({
            'success': True,
            'user': user_data,
            'token': f'fake-jwt-token-{user_data["id"]}'
        })
    else:
        return jsonify({
            'success': False, 
            'error': 'Email o contraseña incorrectos'
        }), 401

@app.route('/api/auth/verify', methods=['GET'])
def verify_token():
    # Endpoint para verificar si el token es válido
    auth_header = request.headers.get('Authorization')
    if auth_header and 'fake-jwt-token' in auth_header:
        return jsonify({'valid': True})
    return jsonify({'valid': False}), 401

@app.route('/api/usuarios', methods=['GET'])
def get_usuarios():
    conn = get_db_connection()
    usuarios = conn.execute('''
        SELECT id, name, email, role, created_at 
        FROM usuarios
    ''').fetchall()
    conn.close()
    return jsonify([dict(usuario) for usuario in usuarios])

# -------------------- RUTAS PARA PRODUCTOS --------------------
@app.route('/api/productos', methods=['GET'])
def get_productos():
    conn = get_db_connection()
    productos = conn.execute('SELECT * FROM productos').fetchall()
    conn.close()
    return jsonify([dict(producto) for producto in productos])

@app.route('/api/productos/<int:id>', methods=['GET'])
def get_producto(id):
    conn = get_db_connection()
    producto = conn.execute('SELECT * FROM productos WHERE id = ?', (id,)).fetchone()
    conn.close()
    return jsonify(dict(producto)) if producto else ('', 404)

@app.route('/api/productos', methods=['POST'])
def agregar_producto():
    data = request.json
    conn = get_db_connection()
    conn.execute('INSERT INTO productos (nombre, tipo, precio, descripcion) VALUES (?, ?, ?, ?)', 
                 (data['nombre'], data['tipo'], data['precio'], data.get('descripcion', '')))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Producto creado'}), 201

@app.route('/api/productos/<int:id>', methods=['PUT'])
def actualizar_producto(id):
    data = request.json
    conn = get_db_connection()
    conn.execute('UPDATE productos SET nombre = ?, tipo = ?, precio = ?, descripcion = ? WHERE id = ?',
                 (data['nombre'], data['tipo'], data['precio'], data.get('descripcion', ''), id))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Producto actualizado'})

@app.route('/api/productos/<int:id>', methods=['DELETE'])
def eliminar_producto(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM productos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Producto eliminado'})

# -------------------- RUTAS PARA CLIENTES --------------------
@app.route('/api/clientes', methods=['GET'])
def get_clientes():
    conn = get_db_connection()
    clientes = conn.execute('SELECT * FROM clientes').fetchall()
    conn.close()
    return jsonify([dict(cliente) for cliente in clientes])

@app.route('/api/clientes/<int:id>', methods=['GET'])
def get_cliente(id):
    conn = get_db_connection()
    cliente = conn.execute('SELECT * FROM clientes WHERE id = ?', (id,)).fetchone()
    conn.close()
    return jsonify(dict(cliente)) if cliente else ('', 404)

@app.route('/api/clientes', methods=['POST'])
def agregar_cliente():
    data = request.json
    conn = get_db_connection()
    conn.execute('INSERT INTO clientes (nombre, email, telefono, direccion, notas) VALUES (?, ?, ?, ?, ?)',
                 (data['nombre'], data.get('email', ''), data.get('telefono', ''), 
                  data.get('direccion', ''), data.get('notas', '')))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Cliente agregado'}), 201

@app.route('/api/clientes/<int:id>', methods=['PUT'])
def actualizar_cliente(id):
    data = request.json
    conn = get_db_connection()
    conn.execute('UPDATE clientes SET nombre = ?, email = ?, telefono = ?, direccion = ?, notas = ? WHERE id = ?',
                 (data['nombre'], data.get('email', ''), data.get('telefono', ''), 
                  data.get('direccion', ''), data.get('notas', ''), id))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Cliente actualizado'})

@app.route('/api/clientes/<int:id>', methods=['DELETE'])
def eliminar_cliente(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM clientes WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Cliente eliminado'})

# -------------------- RUTAS PARA PEDIDOS --------------------
@app.route('/api/pedidos', methods=['GET'])
def get_pedidos():
    conn = get_db_connection()
    pedidos = conn.execute('''
        SELECT p.*, c.nombre as cliente_nombre 
        FROM pedidos p 
        LEFT JOIN clientes c ON p.cliente_id = c.id
    ''').fetchall()
    
    # Obtener productos para cada pedido
    pedidos_con_productos = []
    for pedido in pedidos:
        pedido_dict = dict(pedido)
        productos = conn.execute('''
            SELECT pp.cantidad, pr.nombre, pr.precio, pr.tipo
            FROM pedido_productos pp
            JOIN productos pr ON pp.producto_id = pr.id
            WHERE pp.pedido_id = ?
        ''', (pedido['id'],)).fetchall()
        pedido_dict['productos'] = [dict(producto) for producto in productos]
        pedidos_con_productos.append(pedido_dict)
    
    conn.close()
    return jsonify(pedidos_con_productos)

@app.route('/api/pedidos', methods=['POST'])
def crear_pedido():
    data = request.json
    conn = get_db_connection()
    
    # Crear el pedido
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO pedidos (cliente_id, fecha, encargado_principal, pago_realizado, notas, estado) 
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (
        data.get('cliente_id'),
        data.get('fecha', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        data.get('encargado_principal', ''),
        data.get('pago_realizado', False),
        data.get('notas', ''),
        data.get('estado', 'pendiente')
    ))
    
    pedido_id = cursor.lastrowid
    
    # Agregar productos al pedido
    if 'productos' in data:
        for producto in data['productos']:
            cursor.execute('''
                INSERT INTO pedido_productos (pedido_id, producto_id, cantidad) 
                VALUES (?, ?, ?)
            ''', (pedido_id, producto['producto_id'], producto.get('cantidad', 1)))
    
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Pedido creado', 'id': pedido_id}), 201

@app.route('/api/pedidos/<int:id>', methods=['PUT'])
def actualizar_pedido(id):
    data = request.json
    conn = get_db_connection()
    
    # Actualizar el pedido principal
    conn.execute('''
        UPDATE pedidos 
        SET cliente_id = ?, fecha = ?, encargado_principal = ?, pago_realizado = ?, notas = ?, estado = ?
        WHERE id = ?
    ''', (
        data.get('cliente_id'),
        data.get('fecha'),
        data.get('encargado_principal', ''),
        data.get('pago_realizado', False),
        data.get('notas', ''),
        data.get('estado', 'pendiente'),
        id
    ))
    
    # Actualizar productos si se proporcionan
    if 'productos' in data:
        # Eliminar productos existentes
        conn.execute('DELETE FROM pedido_productos WHERE pedido_id = ?', (id,))
        # Agregar nuevos productos
        for producto in data['productos']:
            conn.execute('''
                INSERT INTO pedido_productos (pedido_id, producto_id, cantidad) 
                VALUES (?, ?, ?)
            ''', (id, producto['producto_id'], producto.get('cantidad', 1)))
    
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Pedido actualizado'})

@app.route('/api/pedidos/<int:id>/estado', methods=['PUT'])
def actualizar_estado_pedido(id):
    data = request.json
    conn = get_db_connection()
    conn.execute('UPDATE pedidos SET estado = ? WHERE id = ?', (data['estado'], id))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Estado del pedido actualizado'})

@app.route('/api/pedidos/<int:id>/pago', methods=['PUT'])
def actualizar_pago_pedido(id):
    data = request.json
    conn = get_db_connection()
    conn.execute('UPDATE pedidos SET pago_realizado = ? WHERE id = ?', (data['pago_realizado'], id))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Estado de pago actualizado'})

@app.route('/api/pedidos/<int:id>', methods=['DELETE'])
def eliminar_pedido(id):
    conn = get_db_connection()
    # Eliminar productos del pedido primero
    conn.execute('DELETE FROM pedido_productos WHERE pedido_id = ?', (id,))
    # Eliminar el pedido
    conn.execute('DELETE FROM pedidos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Pedido eliminado'})

@app.route('/api/pedidos/pendientes', methods=['GET'])
def get_pedidos_pendientes():
    """Obtener pedidos que no tienen venta asociada y están listos para facturar"""
    conn = get_db_connection()
    pedidos = conn.execute('''
        SELECT p.*, c.nombre as cliente_nombre
        FROM pedidos p
        LEFT JOIN clientes c ON p.cliente_id = c.id
        LEFT JOIN ventas v ON p.id = v.pedido_id
        WHERE v.id IS NULL
        AND p.estado = 'completado'
        ORDER BY p.fecha DESC
    ''').fetchall()
    conn.close()
    return jsonify([dict(pedido) for pedido in pedidos])

# -------------------- RUTAS PARA VENTAS --------------------
@app.route('/api/ventas', methods=['GET'])
def get_ventas():
    conn = get_db_connection()
    ventas = conn.execute('''
        SELECT v.*, 
               c.nombre as cliente_nombre, 
               p.nombre as producto_nombre,
               ped.id as pedido_numero
        FROM ventas v 
        LEFT JOIN clientes c ON v.cliente_id = c.id
        LEFT JOIN productos p ON v.producto_id = p.id
        LEFT JOIN pedidos ped ON v.pedido_id = ped.id
        ORDER BY v.id DESC
    ''').fetchall()
    conn.close()
    return jsonify([dict(venta) for venta in ventas])

@app.route('/api/ventas', methods=['POST'])
def registrar_venta():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Si viene pedido_id, obtener datos del pedido
        pedido_id = data.get('pedido_id')
        if pedido_id:
            pedido = cursor.execute('SELECT * FROM pedidos WHERE id = ?', (pedido_id,)).fetchone()
            if not pedido:
                conn.close()
                return jsonify({'error': 'Pedido no encontrado'}), 404
            
            # Usar datos del pedido
            cliente_id = pedido['cliente_id']
            # Para el total, usar el monto del pedido o calcular desde productos
            productos_pedido = cursor.execute('''
                SELECT pp.cantidad, p.precio 
                FROM pedido_productos pp 
                JOIN productos p ON pp.producto_id = p.id 
                WHERE pp.pedido_id = ?
            ''', (pedido_id,)).fetchall()
            
            total = sum(prod['cantidad'] * prod['precio'] for prod in productos_pedido)
        else:
            # Método tradicional - calcular desde producto individual
            producto = cursor.execute('SELECT precio FROM productos WHERE id = ?', 
                                   (data['producto_id'],)).fetchone()
            if not producto:
                conn.close()
                return jsonify({'error': 'Producto no encontrado'}), 404
            
            total = producto['precio'] * data['cantidad']
            cliente_id = data.get('cliente_id')
        
        # Determinar estado de pago
        estado_pago = data.get('estado_pago', 'pendiente')  # Default pendiente
        
        # Crear la venta
        cursor.execute('''
            INSERT INTO ventas (cliente_id, producto_id, cantidad, total, fecha, pedido_id, estado_pago) 
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            cliente_id,
            data.get('producto_id'),  # Puede ser None si viene de pedido con múltiples productos
            data.get('cantidad', 1), 
            total, 
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            pedido_id,
            estado_pago
        ))
        
        venta_id = cursor.lastrowid
        
        # Si la venta está pendiente de pago, crear cuenta por cobrar automáticamente
        if estado_pago == 'pendiente':
            numero_factura = f"FAC-{venta_id:04d}"
            fecha_vencimiento = datetime.now().date() + timedelta(days=30)  # 30 días para pagar
            
            cursor.execute('''
                INSERT INTO cuentas_por_cobrar 
                (numero_factura, cliente_id, venta_id, pedido_id, monto, saldo, fecha_vencimiento, estado)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                numero_factura,
                cliente_id,
                venta_id,
                pedido_id,
                total,
                total,  # saldo inicial = monto total
                fecha_vencimiento.strftime("%Y-%m-%d"),
                'pendiente'
            ))
        
        # Si hay pedido_id, actualizar estado del pedido
        if pedido_id and estado_pago == 'pagado':
            cursor.execute('UPDATE pedidos SET estado_pago = ? WHERE id = ?', ('pagado', pedido_id))
        
        conn.commit()
        conn.close()
        return jsonify({
            'mensaje': 'Venta registrada',
            'venta_id': venta_id,
            'cuenta_por_cobrar_generada': estado_pago == 'pendiente'
        }), 201
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/ventas/<int:id>', methods=['DELETE'])
def eliminar_venta(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM ventas WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Venta eliminada'})

# -------------------- RUTAS PARA CUENTAS POR COBRAR --------------------
@app.route('/api/cuentas-por-cobrar', methods=['GET'])
def get_cuentas_por_cobrar():
    conn = get_db_connection()
    cuentas = conn.execute('''
        SELECT c.*, 
               cl.nombre as cliente_nombre,
               p.id as pedido_numero
        FROM cuentas_por_cobrar c
        LEFT JOIN clientes cl ON c.cliente_id = cl.id
        LEFT JOIN pedidos p ON c.pedido_id = p.id
        ORDER BY c.fecha_vencimiento ASC
    ''').fetchall()
    
    # Calcular días vencidos para cada cuenta
    cuentas_con_datos = []
    for cuenta in cuentas:
        cuenta_dict = dict(cuenta)
        
        # Calcular días vencidos
        from datetime import datetime, date
        if cuenta['fecha_vencimiento']:
            try:
                fecha_venc = datetime.strptime(cuenta['fecha_vencimiento'], '%Y-%m-%d').date()
                hoy = date.today()
                dias_diff = (hoy - fecha_venc).days
                cuenta_dict['dias_vencido'] = max(0, dias_diff)
                
                # Actualizar estado según días vencidos
                if cuenta_dict['estado'] == 'pendiente' and dias_diff > 0:
                    cuenta_dict['estado'] = 'vencido'
                elif cuenta_dict['estado'] == 'vencido' and dias_diff <= 0:
                    cuenta_dict['estado'] = 'pendiente'
            except:
                cuenta_dict['dias_vencido'] = 0
        
        cuentas_con_datos.append(cuenta_dict)
    
    conn.close()
    return jsonify(cuentas_con_datos)

@app.route('/api/cuentas-por-cobrar', methods=['POST'])
def crear_cuenta_por_cobrar():
    data = request.json
    conn = get_db_connection()
    
    # Calcular saldo inicial (monto - monto_pagado)
    monto = data['monto']
    monto_pagado = data.get('monto_pagado', 0)
    saldo = monto - monto_pagado
    
    # Determinar estado inicial
    estado = 'pagado' if saldo <= 0 else data.get('estado', 'pendiente')
    
    conn.execute('''
        INSERT INTO cuentas_por_cobrar 
        (numero_factura, cliente_id, pedido_id, monto, monto_pagado, saldo, 
         fecha_vencimiento, estado, notas) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['numero_factura'],
        data['cliente_id'],
        data.get('pedido_id'),
        monto,
        monto_pagado,
        saldo,
        data['fecha_vencimiento'],
        estado,
        data.get('notas', '')
    ))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Cuenta por cobrar creada'}), 201

@app.route('/api/cuentas-por-cobrar/<int:id>', methods=['PUT'])
def actualizar_cuenta_por_cobrar(id):
    data = request.json
    conn = get_db_connection()
    
    # Si es una actualización de pago
    if 'monto_pagado' in data:
        # Obtener datos actuales
        cuenta_actual = conn.execute('SELECT * FROM cuentas_por_cobrar WHERE id = ?', (id,)).fetchone()
        if not cuenta_actual:
            conn.close()
            return jsonify({'error': 'Cuenta no encontrada'}), 404
        
        # Calcular nuevo saldo
        monto = cuenta_actual['monto']
        nuevo_monto_pagado = data['monto_pagado']
        nuevo_saldo = monto - nuevo_monto_pagado
        
        # Actualizar estado según saldo
        nuevo_estado = 'pagado' if nuevo_saldo <= 0 else data.get('estado', cuenta_actual['estado'])
        
        conn.execute('''
            UPDATE cuentas_por_cobrar 
            SET monto_pagado = ?, saldo = ?, estado = ?, notas = ?
            WHERE id = ?
        ''', (nuevo_monto_pagado, nuevo_saldo, nuevo_estado, data.get('notas', cuenta_actual['notas']), id))
    else:
        # Actualización completa
        monto = data.get('monto')
        monto_pagado = data.get('monto_pagado', 0)
        saldo = monto - monto_pagado
        estado = 'pagado' if saldo <= 0 else data.get('estado', 'pendiente')
        
        conn.execute('''
            UPDATE cuentas_por_cobrar 
            SET numero_factura = ?, cliente_id = ?, pedido_id = ?, monto = ?, 
                monto_pagado = ?, saldo = ?, fecha_vencimiento = ?, estado = ?, notas = ?
            WHERE id = ?
        ''', (
            data['numero_factura'], data['cliente_id'], data.get('pedido_id'),
            monto, monto_pagado, saldo, data['fecha_vencimiento'], estado,
            data.get('notas', ''), id
        ))
    
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Cuenta por cobrar actualizada'})

@app.route('/api/cuentas-por-cobrar/<int:id>', methods=['DELETE'])
def eliminar_cuenta_por_cobrar(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM cuentas_por_cobrar WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Cuenta por cobrar eliminada'})

@app.route('/api/cuentas-por-cobrar/stats', methods=['GET'])
def estadisticas_cuentas_por_cobrar():
    conn = get_db_connection()
    
    # Total por cobrar
    total_result = conn.execute('SELECT SUM(saldo) FROM cuentas_por_cobrar WHERE estado != "pagado"').fetchone()
    total_por_cobrar = total_result[0] or 0
    
    # Facturas pendientes
    pendientes_result = conn.execute('SELECT COUNT(*) FROM cuentas_por_cobrar WHERE estado = "pendiente"').fetchone()
    facturas_pendientes = pendientes_result[0] or 0
    
    # Facturas vencidas
    vencidas_result = conn.execute('SELECT COUNT(*) FROM cuentas_por_cobrar WHERE estado = "vencido"').fetchone()
    facturas_vencidas = vencidas_result[0] or 0
    
    # Total de facturas
    total_facturas_result = conn.execute('SELECT COUNT(*) FROM cuentas_por_cobrar').fetchone()
    total_facturas = total_facturas_result[0] or 0
    
    conn.close()
    
    return jsonify({
        'total_por_cobrar': total_por_cobrar,
        'facturas_pendientes': facturas_pendientes,
        'facturas_vencidas': facturas_vencidas,
        'total_facturas': total_facturas
    })

@app.route('/api/cuentas-por-cobrar/<int:id>/marcar-pagado', methods=['PUT'])
def marcar_cuenta_como_pagada(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener cuenta actual
        cuenta = cursor.execute('SELECT * FROM cuentas_por_cobrar WHERE id = ?', (id,)).fetchone()
        if not cuenta:
            conn.close()
            return jsonify({'error': 'Cuenta no encontrada'}), 404
        
        # Marcar cuenta como completamente pagada
        cursor.execute('''
            UPDATE cuentas_por_cobrar 
            SET monto_pagado = monto, saldo = 0, estado = 'pagado'
            WHERE id = ?
        ''', (id,))
        
        # Actualizar la venta relacionada como pagada
        if cuenta['venta_id']:
            cursor.execute('''
                UPDATE ventas 
                SET estado_pago = 'pagado'
                WHERE id = ?
            ''', (cuenta['venta_id'],))
        
        # Actualizar el pedido relacionado como pagado
        if cuenta['pedido_id']:
            cursor.execute('''
                UPDATE pedidos 
                SET estado_pago = 'pagado'
                WHERE id = ?
            ''', (cuenta['pedido_id'],))
        
        conn.commit()
        conn.close()
        return jsonify({
            'mensaje': 'Cuenta marcada como pagada',
            'venta_actualizada': bool(cuenta['venta_id']),
            'pedido_actualizado': bool(cuenta['pedido_id'])
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

# -------------------- RUTAS PARA CUENTAS POR PAGAR --------------------
@app.route('/api/cuentas-por-pagar', methods=['GET'])
def get_cuentas_por_pagar():
    conn = get_db_connection()
    cuentas = conn.execute('''
        SELECT *
        FROM cuentas_por_pagar
        ORDER BY fecha_vencimiento ASC
    ''').fetchall()
    
    # Calcular días vencidos para cada cuenta
    cuentas_con_datos = []
    for cuenta in cuentas:
        cuenta_dict = dict(cuenta)
        
        # Calcular días vencidos
        from datetime import datetime, date
        if cuenta['fecha_vencimiento']:
            try:
                fecha_venc = datetime.strptime(cuenta['fecha_vencimiento'], '%Y-%m-%d').date()
                hoy = date.today()
                dias_diff = (hoy - fecha_venc).days
                cuenta_dict['dias_vencido'] = max(0, dias_diff)
                
                # Actualizar estado según días vencidos
                if cuenta_dict['estado'] == 'pendiente' and dias_diff > 0:
                    cuenta_dict['estado'] = 'vencido'
                elif cuenta_dict['estado'] == 'vencido' and dias_diff <= 0:
                    cuenta_dict['estado'] = 'pendiente'
            except:
                cuenta_dict['dias_vencido'] = 0
        
        cuentas_con_datos.append(cuenta_dict)
    
    conn.close()
    return jsonify(cuentas_con_datos)

@app.route('/api/cuentas-por-pagar', methods=['POST'])
def crear_cuenta_por_pagar():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Auto-generar codigo_factura: BILL001, BILL002, etc
        ultimo_codigo = cursor.execute('''
            SELECT codigo_factura FROM cuentas_por_pagar 
            WHERE codigo_factura LIKE 'BILL%' 
            ORDER BY id DESC LIMIT 1
        ''').fetchone()
        
        if ultimo_codigo:
            # Extraer numero y sumar 1
            numero_str = ultimo_codigo[0].replace('BILL', '')
            try:
                numero = int(numero_str) + 1
            except:
                numero = 1
        else:
            numero = 1
        
        codigo_factura = f"BILL{numero:03d}"
        
        # Calcular saldo inicial (monto - monto_pagado)
        monto = data['monto']
        monto_pagado = data.get('monto_pagado', 0)
        saldo = monto - monto_pagado
        
        # Determinar estado inicial
        estado = 'pagado' if saldo <= 0 else data.get('estado', 'pendiente')
        
        cursor.execute('''
            INSERT INTO cuentas_por_pagar 
            (codigo_factura, proveedor, monto, monto_pagado, saldo, 
             fecha_vencimiento, estado, descripcion) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            codigo_factura,
            data['proveedor'],
            monto,
            monto_pagado,
            saldo,
            data['fecha_vencimiento'],
            estado,
            data.get('descripcion', '')
        ))
        
        conn.commit()
        conn.close()
        return jsonify({
            'mensaje': 'Cuenta por pagar creada',
            'codigo_factura': codigo_factura
        }), 201
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cuentas-por-pagar/<int:id>', methods=['PUT'])
def actualizar_cuenta_por_pagar(id):
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Si es una actualización de pago
        if 'monto_pagado' in data:
            # Obtener datos actuales
            cuenta_actual = cursor.execute('SELECT * FROM cuentas_por_pagar WHERE id = ?', (id,)).fetchone()
            if not cuenta_actual:
                conn.close()
                return jsonify({'error': 'Cuenta no encontrada'}), 404
            
            # Calcular nuevo saldo
            monto = cuenta_actual['monto']
            nuevo_monto_pagado = data['monto_pagado']
            nuevo_saldo = monto - nuevo_monto_pagado
            
            # Actualizar estado según saldo
            nuevo_estado = 'pagado' if nuevo_saldo <= 0 else data.get('estado', cuenta_actual['estado'])
            fecha_pago = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if nuevo_saldo <= 0 else cuenta_actual['fecha_pago']
            
            cursor.execute('''
                UPDATE cuentas_por_pagar 
                SET monto_pagado = ?, saldo = ?, estado = ?, fecha_pago = ?, descripcion = ?
                WHERE id = ?
            ''', (nuevo_monto_pagado, nuevo_saldo, nuevo_estado, fecha_pago, 
                  data.get('descripcion', cuenta_actual['descripcion']), id))
        else:
            # Actualización completa
            monto = data.get('monto')
            monto_pagado = data.get('monto_pagado', 0)
            saldo = monto - monto_pagado
            estado = 'pagado' if saldo <= 0 else data.get('estado', 'pendiente')
            
            cursor.execute('''
                UPDATE cuentas_por_pagar 
                SET proveedor = ?, monto = ?, monto_pagado = ?, saldo = ?, 
                    fecha_vencimiento = ?, estado = ?, descripcion = ?
                WHERE id = ?
            ''', (
                data['proveedor'], monto, monto_pagado, saldo, 
                data['fecha_vencimiento'], estado, data.get('descripcion', ''), id
            ))
        
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Cuenta por pagar actualizada'})
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cuentas-por-pagar/<int:id>', methods=['DELETE'])
def eliminar_cuenta_por_pagar(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM cuentas_por_pagar WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'mensaje': 'Cuenta por pagar eliminada'})

@app.route('/api/cuentas-por-pagar/stats', methods=['GET'])
def estadisticas_cuentas_por_pagar():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Total por pagar
        total_result = cursor.execute('SELECT SUM(saldo) FROM cuentas_por_pagar WHERE estado != "pagado"').fetchone()
        total_por_pagar = total_result[0] or 0
        
        # Facturas pendientes
        pendientes_result = cursor.execute('SELECT COUNT(*) FROM cuentas_por_pagar WHERE estado = "pendiente"').fetchone()
        facturas_pendientes = pendientes_result[0] or 0
        
        # Facturas vencidas
        vencidas_result = cursor.execute('SELECT COUNT(*) FROM cuentas_por_pagar WHERE estado = "vencido"').fetchone()
        facturas_vencidas = vencidas_result[0] or 0
        
        # Próximas a vencer (7 días)
        from datetime import date, timedelta
        fecha_limite = (date.today() + timedelta(days=7)).strftime('%Y-%m-%d')
        proximas_result = cursor.execute('''
            SELECT COUNT(*) FROM cuentas_por_pagar 
            WHERE estado = "pendiente" AND fecha_vencimiento <= ?
        ''', (fecha_limite,)).fetchone()
        proximas_vencer = proximas_result[0] or 0
        
        conn.close()
        
        return jsonify({
            'total_por_pagar': total_por_pagar,
            'facturas_pendientes': facturas_pendientes,
            'facturas_vencidas': facturas_vencidas,
            'proximas_vencer': proximas_vencer
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cuentas-por-pagar/<int:id>/marcar-pagado', methods=['PUT'])
def marcar_cuenta_por_pagar_como_pagada(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener cuenta actual
        cuenta = cursor.execute('SELECT * FROM cuentas_por_pagar WHERE id = ?', (id,)).fetchone()
        if not cuenta:
            conn.close()
            return jsonify({'error': 'Cuenta no encontrada'}), 404
        
        # Marcar como completamente pagada
        cursor.execute('''
            UPDATE cuentas_por_pagar 
            SET monto_pagado = monto, saldo = 0, estado = 'pagado', fecha_pago = ?
            WHERE id = ?
        ''', (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), id))
        
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Cuenta marcada como pagada'})
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cuentas-por-cobrar/all', methods=['DELETE'])
def delete_all_receivables():
    """Eliminar todas las cuentas por cobrar"""
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM cuentas_por_cobrar')
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Todas las cuentas por cobrar han sido eliminadas'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cuentas-por-pagar/all', methods=['DELETE'])
def delete_all_payables():
    """Eliminar todas las cuentas por pagar"""
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM cuentas_por_pagar')
        conn.commit()
        conn.close()
        return jsonify({'mensaje': 'Todas las cuentas por pagar han sido eliminadas'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/restore-productos-originales', methods=['POST'])
def restore_productos_originales():
    """
    URGENTE - Restaurar productos reales eliminados accidentalmente
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # PRIMERO: Borrar productos actuales (ejemplo/demo)
        cursor.execute('DELETE FROM productos')
        
        # SEGUNDO: Resetear secuencia productos
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="productos"')
        
        # TERCERO: Insertar productos reales originales
        productos_originales = [
            ("SCENE ANIMATION", "vfx", 2079.20, "Animación de escena completa"),
            ("SCENE", "vfx", 1315.06, "Escena de video profesional"),
            ("ANIMATED 2.0 FRAME", "vfx", 805.60, "Frame animado versión 2.0"),
            ("TRANSITION", "vfx", 725.67, "Transición de video profesional"),
            ("INTRO", "vfx", 275.84, "Introducción animada"),
            ("LOGO ANIMATION", "vfx", 215.91, "Animación de logo"),
            ("POST (1 SLIDE)", "gfx", 69.93, "Post de 1 slide para redes sociales"),
            ("ANIMATED OUTRO", "vfx", 65.98, "Outro animado"),
            ("POST RAIMATION", "gfx", 31.38, "Post con animación básica"),
            ("2.0 FRAME", "gfx", 27.98, "Frame versión 2.0"),
            ("LOWERTHIRD", "gfx", 26.87, "Lower third gráfico")
        ]
        
        cursor.executemany('''
            INSERT INTO productos (nombre, tipo, precio, descripcion) 
            VALUES (?, ?, ?, ?)
        ''', productos_originales)
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Productos originales restaurados exitosamente',
            'productos_restaurados': len(productos_originales),
            'total_productos': 11,
            'productos': [p[0] for p in productos_originales]
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': f'Error al restaurar productos: {str(e)}'}), 500

@app.route('/api/reset-database', methods=['POST'])
def reset_database_for_production():
    """
    SOLO PARA PRODUCCION - Resetea completamente la base de datos
    Borra todos los datos y reinicia secuencias en 1
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # BORRAR TODOS LOS DATOS
        cursor.execute('DELETE FROM ventas')
        cursor.execute('DELETE FROM pedido_productos')
        cursor.execute('DELETE FROM pedidos') 
        cursor.execute('DELETE FROM clientes')
        cursor.execute('DELETE FROM productos')
        cursor.execute('DELETE FROM cuentas_por_cobrar')
        cursor.execute('DELETE FROM cuentas_por_pagar')
        
        # RESETEAR SECUENCIAS AUTOINCREMENT
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="ventas"')
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="pedidos"')
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="pedido_productos"')
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="clientes"')
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="productos"')
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="cuentas_por_cobrar"')
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="cuentas_por_pagar"')
        
        # INSERTAR DATOS EJEMPLO PARA DEMO
        # Productos ejemplo
        cursor.execute('''
            INSERT INTO productos (nombre, tipo, precio, descripcion) VALUES
            ('SCENE ANIMATION', 'vfx', 2079.20, 'Animación de escena completa con efectos profesionales'),
            ('LOGO DESIGN', 'gfx', 500.00, 'Diseño de logo corporativo profesional'),
            ('VIDEO EDIT PRO', 'vfx', 1200.00, 'Edición profesional de video con efectos'),
            ('BRANDING PACKAGE', 'gfx', 800.00, 'Paquete completo de identidad corporativa'),
            ('3D MODELING', 'vfx', 1500.00, 'Modelado 3D para animación y renders')
        ''')
        
        # Clientes ejemplo  
        cursor.execute('''
            INSERT INTO clientes (nombre, email, telefono, direccion, notas) VALUES
            ('Empresa Innovadora S.A.', 'contacto@innovadora.com', '+1 (555) 123-4567', 'Av. Tecnología 123, Centro Empresarial', 'Cliente corporativo - Proyectos grandes'),
            ('Estudio Creativo Luna', 'info@estudioluna.com', '+1 (555) 987-6543', 'Calle Arte 456, Distrito Creativo', 'Estudio de diseño - Colaboraciones frecuentes'),
            ('Digital Marketing Pro', 'hello@digitalmarketing.com', '+1 (555) 555-0199', 'Torre Comercial 789, Piso 15', 'Agencia de marketing - Campañas mensuales')
        ''')
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'mensaje': 'Base de datos reseteada para producción exitosamente',
            'productos_creados': 5,
            'clientes_creados': 3,
            'secuencias_reiniciadas': ['clientes', 'productos', 'pedidos', 'ventas', 'cuentas_por_cobrar', 'cuentas_por_pagar']
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': f'Error al resetear base de datos: {str(e)}'}), 500

# -------------------- RUTA DE INICIALIZACIÓN --------------------
@app.route('/api/init-db', methods=['POST'])
def initialize_database():
    init_db()
    return jsonify({'mensaje': 'Base de datos inicializada correctamente'})


@app.route('/api/reset-sequences', methods=['POST'])
def reset_sequences():
    """Reiniciar secuencias de ID cuando las tablas están vacías"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener datos del request para saber qué tablas resetear
        data = request.json if request.json else {}
        force_reset = data.get('force', False)
        specific_tables = data.get('tables', [])
        
        # Lista de todas las tablas con autoincrement
        all_tables = ['productos', 'clientes', 'pedidos', 'ventas', 'pedido_productos']
        tables_to_check = specific_tables if specific_tables else all_tables
        
        tables_to_reset = []
        tables_with_data = []
        
        # Verificar cada tabla
        for table in tables_to_check:
            count = cursor.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
            if count == 0:
                tables_to_reset.append(table)
            else:
                tables_with_data.append(table)
        
        # Si force_reset es True, resetear también tablas con datos (PELIGROSO)
        if force_reset and tables_with_data:
            tables_to_reset.extend(tables_with_data)
        
        # Resetear secuencias
        for table in tables_to_reset:
            cursor.execute(f"DELETE FROM sqlite_sequence WHERE name='{table}'")
        
        conn.commit()
        conn.close()
        
        result = {
            'mensaje': f'Secuencias reseteadas para: {", ".join(tables_to_reset) if tables_to_reset else "ninguna tabla"}',
            'tablas_reseteadas': tables_to_reset,
            'tablas_con_datos': tables_with_data
        }
        
        if tables_with_data and not force_reset:
            result['advertencia'] = f'Las siguientes tablas tienen datos y no fueron reseteadas: {", ".join(tables_with_data)}. Usa force=true para forzar el reseteo.'
        
        return jsonify(result)
    
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/reset-sequences/<string:table_name>', methods=['POST'])
def reset_single_sequence(table_name):
    """Reiniciar secuencia de una tabla específica"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Validar nombre de tabla
    valid_tables = ['productos', 'clientes', 'pedidos', 'ventas', 'pedido_productos']
    if table_name not in valid_tables:
        conn.close()
        return jsonify({'error': f'Tabla inválida. Tablas válidas: {", ".join(valid_tables)}'}), 400
    
    try:
        data = request.json if request.json else {}
        force = data.get('force', False)
        
        # Verificar si la tabla tiene datos
        count = cursor.execute(f'SELECT COUNT(*) FROM {table_name}').fetchone()[0]
        
        if count > 0 and not force:
            conn.close()
            return jsonify({
                'error': f'La tabla {table_name} tiene {count} registros. Usa force=true para forzar el reseteo.',
                'registros': count
            }), 400
        
        # Resetear la secuencia
        cursor.execute(f"DELETE FROM sqlite_sequence WHERE name='{table_name}'")
        conn.commit()
        conn.close()
        
        return jsonify({
            'mensaje': f'Secuencia de {table_name} reseteada correctamente',
            'tabla': table_name,
            'registros_antes': count,
            'forzado': force
        })
    
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/sequences/status', methods=['GET'])
def get_sequences_status():
    """Obtener estado actual de todas las secuencias"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener todas las secuencias
        sequences = cursor.execute('SELECT name, seq FROM sqlite_sequence ORDER BY name').fetchall()
        
        # Obtener conteo de registros por tabla
        status = []
        for seq_name, seq_value in sequences:
            count = cursor.execute(f'SELECT COUNT(*) FROM {seq_name}').fetchone()[0]
            max_id = cursor.execute(f'SELECT MAX(id) FROM {seq_name}').fetchone()[0] or 0
            
            status.append({
                'tabla': seq_name,
                'secuencia_actual': seq_value,
                'registros_actuales': count,
                'id_maximo': max_id,
                'necesita_reset': count == 0 and seq_value > 0
            })
        
        conn.close()
        return jsonify({
            'secuencias': status,
            'tablas_vacias_con_secuencia': [s for s in status if s['necesita_reset']]
        })
    
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

# -------------------- FUNCIONES HELPER PARA REPORTES --------------------
def get_periodo_fechas(periodo):
    """Calcula fechas de inicio y fin según el periodo solicitado"""
    hoy = datetime.now()
    
    if periodo == 'semana':
        inicio = hoy - timedelta(days=7)
        fin = hoy
        periodo_anterior_inicio = inicio - timedelta(days=7)
        periodo_anterior_fin = inicio
    elif periodo == 'mes':
        inicio = hoy.replace(day=1)
        fin = hoy
        if inicio.month == 1:
            periodo_anterior_inicio = inicio.replace(year=inicio.year-1, month=12, day=1)
            periodo_anterior_fin = inicio.replace(year=inicio.year-1, month=12, day=31)
        else:
            periodo_anterior_inicio = inicio.replace(month=inicio.month-1, day=1)
            if inicio.month-1 in [1,3,5,7,8,10,12]:
                periodo_anterior_fin = inicio.replace(month=inicio.month-1, day=31)
            elif inicio.month-1 in [4,6,9,11]:
                periodo_anterior_fin = inicio.replace(month=inicio.month-1, day=30)
            else:
                periodo_anterior_fin = inicio.replace(month=inicio.month-1, day=28)
    elif periodo == 'trimestre':
        inicio = hoy - timedelta(days=90)
        fin = hoy
        periodo_anterior_inicio = inicio - timedelta(days=90)
        periodo_anterior_fin = inicio
    elif periodo == 'ano':
        inicio = hoy.replace(month=1, day=1)
        fin = hoy
        periodo_anterior_inicio = inicio.replace(year=inicio.year-1)
        periodo_anterior_fin = inicio.replace(year=inicio.year-1, month=12, day=31)
    else:
        inicio = hoy.replace(day=1)
        fin = hoy
        periodo_anterior_inicio = inicio.replace(month=inicio.month-1, day=1)
        periodo_anterior_fin = inicio.replace(month=inicio.month-1, day=31)
    
    return inicio.strftime('%Y-%m-%d'), fin.strftime('%Y-%m-%d'), periodo_anterior_inicio.strftime('%Y-%m-%d'), periodo_anterior_fin.strftime('%Y-%m-%d')

def calcular_crecimiento(actual, anterior):
    """Calcula el porcentaje de crecimiento entre dos periodos"""
    if anterior == 0:
        return 100.0 if actual > 0 else 0.0
    return round(((actual - anterior) / anterior) * 100, 1)

# -------------------- ENDPOINTS DE REPORTES --------------------
@app.route('/api/reportes/dashboard', methods=['GET'])
def get_reporte_dashboard():
    """Estadisticas para modulo reportes"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Ventas totales de TODA la tabla (sin filtro de fecha)
        cursor.execute('SELECT COALESCE(SUM(total), 0) FROM ventas')
        ventas_totales = cursor.fetchone()[0]
        
        # Total pedidos de TODA la tabla
        cursor.execute('SELECT COUNT(*) FROM pedidos')
        total_pedidos = cursor.fetchone()[0]
        
        # Valor promedio
        if total_pedidos > 0:
            valor_promedio = ventas_totales / total_pedidos
        else:
            valor_promedio = 0
        
        # Nuevos clientes (todos los clientes para testing)
        cursor.execute('SELECT COUNT(*) FROM clientes')
        nuevos_clientes = cursor.fetchone()[0]
        
        conn.close()
        
        result = {
            'ventas_totales': float(ventas_totales),
            'total_pedidos': total_pedidos,
            'valor_promedio': float(valor_promedio),
            'nuevos_clientes': nuevos_clientes,
            'crecimiento_ventas': 0,
            'crecimiento_pedidos': 0,
            'crecimiento_promedio': 0,
            'crecimiento_clientes': 0
        }
        
        print(f"📊 Reportes stats calculadas: {result}")
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Error en reportes stats: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/ingresos-tipo', methods=['GET'])
def get_ingresos_tipo():
    """Endpoint para ingresos por tipo GFX/VFX"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # PRIMERO: Intentar con VENTAS
        cursor.execute('''
            SELECT 
                UPPER(p.tipo) as tipo,
                COALESCE(SUM(v.total), 0) as total_ingresos,
                COUNT(v.id) as cantidad
            FROM ventas v
            JOIN productos p ON v.producto_id = p.id
            GROUP BY UPPER(p.tipo)
            ORDER BY total_ingresos DESC
        ''')
        ingresos_data = cursor.fetchall()
        
        # FALLBACK: Si no hay ventas, usar PEDIDOS
        if not ingresos_data:
            cursor.execute('''
                SELECT 
                    UPPER(p.tipo) as tipo,
                    COALESCE(SUM(pe.total), 0) as total_ingresos,
                    COUNT(pe.id) as cantidad
                FROM pedidos pe
                JOIN productos p ON pe.producto_id = p.id
                GROUP BY UPPER(p.tipo)
                ORDER BY total_ingresos DESC
            ''')
            ingresos_data = cursor.fetchall()
        
        # Calcular total general
        total_general = sum(row[1] for row in ingresos_data) if ingresos_data else 1
        
        resultado = {}
        for row in ingresos_data:
            tipo = row[0]  # Ya viene en UPPER desde la query
            total = float(row[1])
            cantidad = row[2]
            porcentaje = round((total / total_general * 100), 1) if total_general > 0 else 0
            resultado[tipo] = {
                'total': round(total, 2),
                'porcentaje': porcentaje,
                'cantidad': cantidad
            }
        
        # Asegurar que siempre tengamos GFX y VFX
        if 'GFX' not in resultado:
            resultado['GFX'] = {'total': 0, 'porcentaje': 0, 'cantidad': 0}
        if 'VFX' not in resultado:
            resultado['VFX'] = {'total': 0, 'porcentaje': 0, 'cantidad': 0}
        
        conn.close()
        print(f"💰 Ingresos por tipo calculados: {resultado}")
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Error en ingresos-tipo: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/tendencia', methods=['GET'])
def get_tendencia():
    """Endpoint para tendencia temporal"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        periodo = request.args.get('periodo', 'mes')
        
        # Simplificamos para mostrar datos sin filtros de fecha complicados
        if periodo == 'semana':
            # Agrupar por fecha simple
            query = '''
                SELECT 
                    DATE(v.fecha) as fecha,
                    COALESCE(SUM(CASE WHEN UPPER(p.tipo) = 'VFX' THEN v.total ELSE 0 END), 0) as vfx,
                    COALESCE(SUM(CASE WHEN UPPER(p.tipo) = 'GFX' THEN v.total ELSE 0 END), 0) as gfx,
                    COALESCE(SUM(v.total), 0) as total
                FROM ventas v
                JOIN productos p ON v.producto_id = p.id
                GROUP BY DATE(v.fecha)
                ORDER BY DATE(v.fecha) DESC
                LIMIT 7
            '''
        else:
            # Agrupar por mes (formato simple)
            query = '''
                SELECT 
                    strftime('%Y-%m', v.fecha) as mes,
                    COALESCE(SUM(CASE WHEN UPPER(p.tipo) = 'VFX' THEN v.total ELSE 0 END), 0) as vfx,
                    COALESCE(SUM(CASE WHEN UPPER(p.tipo) = 'GFX' THEN v.total ELSE 0 END), 0) as gfx,
                    COALESCE(SUM(v.total), 0) as total
                FROM ventas v
                JOIN productos p ON v.producto_id = p.id
                GROUP BY strftime('%Y-%m', v.fecha)
                ORDER BY strftime('%Y-%m', v.fecha) DESC
                LIMIT 6
            '''
        
        cursor.execute(query)
        resultados = cursor.fetchall()
        
        # Si no hay ventas, usar pedidos como fallback
        if not resultados:
            if periodo == 'semana':
                query_pedidos = '''
                    SELECT 
                        DATE(p.fecha) as fecha,
                        COALESCE(SUM(CASE WHEN UPPER(pr.tipo) = 'VFX' THEN pp.assigned_payment ELSE 0 END), 0) as vfx,
                        COALESCE(SUM(CASE WHEN UPPER(pr.tipo) = 'GFX' THEN pp.assigned_payment ELSE 0 END), 0) as gfx,
                        COALESCE(SUM(pp.assigned_payment), 0) as total
                    FROM pedidos p
                    JOIN pedido_productos pp ON p.id = pp.pedido_id
                    JOIN productos pr ON pp.producto_id = pr.id
                    GROUP BY DATE(p.fecha)
                    ORDER BY DATE(p.fecha) DESC
                    LIMIT 7
                '''
            else:
                query_pedidos = '''
                    SELECT 
                        strftime('%Y-%m', p.fecha) as mes,
                        COALESCE(SUM(CASE WHEN UPPER(pr.tipo) = 'VFX' THEN pp.assigned_payment ELSE 0 END), 0) as vfx,
                        COALESCE(SUM(CASE WHEN UPPER(pr.tipo) = 'GFX' THEN pp.assigned_payment ELSE 0 END), 0) as gfx,
                        COALESCE(SUM(pp.assigned_payment), 0) as total
                    FROM pedidos p
                    JOIN pedido_productos pp ON p.id = pp.pedido_id
                    JOIN productos pr ON pp.producto_id = pr.id
                    GROUP BY strftime('%Y-%m', p.fecha)
                    ORDER BY strftime('%Y-%m', p.fecha) DESC
                    LIMIT 6
                '''
            cursor.execute(query_pedidos)
            resultados = cursor.fetchall()
        
        # Formatear resultados
        tendencia = []
        for i, row in enumerate(resultados):
            if periodo == 'semana':
                periodo_nombre = f"Día {i+1}" if row[0] else f"Día {i+1}"
            else:
                # Convertir YYYY-MM a nombre más legible
                if row[0]:
                    try:
                        año, mes = row[0].split('-')
                        meses = {
                            '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
                            '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
                            '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
                        }
                        mes_nombre = meses.get(mes, f'Mes {mes}')
                        periodo_nombre = f'{mes_nombre} {año}'
                    except:
                        periodo_nombre = f'Periodo {i+1}'
                else:
                    periodo_nombre = f'Periodo {i+1}'
            
            tendencia.append({
                'periodo': periodo_nombre,
                'vfx': float(row[1]) if row[1] else 0,
                'gfx': float(row[2]) if row[2] else 0,
                'total': float(row[3]) if row[3] else 0
            })
        
        conn.close()
        print(f"📈 Tendencia calculada: {len(tendencia)} periodos")
        return jsonify(tendencia)
        
    except Exception as e:
        print(f"❌ Error en tendencia: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/productos-top', methods=['GET'])
def get_productos_top():
    """Endpoint para productos más vendidos"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # PRIMERO: Intentar con VENTAS (datos reales)
        cursor.execute('''
            SELECT 
                p.nombre,
                p.tipo,
                COUNT(v.id) as pedidos,
                COALESCE(SUM(v.total), 0) as ingresos
            FROM ventas v
            JOIN productos p ON v.producto_id = p.id
            GROUP BY p.id, p.nombre, p.tipo
            ORDER BY ingresos DESC
            LIMIT 10
        ''')
        productos_data = cursor.fetchall()
        
        # FALLBACK: Si no hay ventas, usar PEDIDOS
        if not productos_data:
            cursor.execute('''
                SELECT 
                    p.nombre,
                    p.tipo,
                    COUNT(pe.id) as pedidos,
                    COALESCE(SUM(pe.total), 0) as ingresos
                FROM pedidos pe
                JOIN productos p ON pe.producto_id = p.id
                GROUP BY p.id, p.nombre, p.tipo
                ORDER BY ingresos DESC
                LIMIT 10
            ''')
            productos_data = cursor.fetchall()
        
        resultado = []
        for row in productos_data:
            promedio = row[3] / row[2] if row[2] > 0 else 0
            resultado.append({
                'nombre': row[0],
                'tipo': row[1].upper() if row[1] else 'N/A',
                'pedidos': row[2],
                'ingresos': round(float(row[3]), 2),
                'promedio': round(float(promedio), 2)
            })
        
        conn.close()
        print(f"📊 Productos más vendidos encontrados: {len(resultado)}")
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Error en productos-top: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/clientes-top', methods=['GET'])
def get_clientes_top():
    """Endpoint para mejores clientes"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # PRIMERO: Intentar con VENTAS (datos reales)
        cursor.execute('''
            SELECT 
                c.nombre,
                COUNT(v.id) as pedidos,
                COALESCE(SUM(v.total), 0) as ingresos,
                MAX(v.fecha) as ultimo_pedido
            FROM ventas v
            JOIN clientes c ON v.cliente_id = c.id
            GROUP BY c.id, c.nombre
            ORDER BY ingresos DESC
            LIMIT 10
        ''')
        clientes_data = cursor.fetchall()
        
        # FALLBACK: Si no hay ventas, usar PEDIDOS
        if not clientes_data:
            cursor.execute('''
                SELECT 
                    c.nombre,
                    COUNT(pe.id) as pedidos,
                    COALESCE(SUM(pe.total), 0) as ingresos,
                    MAX(pe.fecha) as ultimo_pedido
                FROM pedidos pe
                JOIN clientes c ON pe.cliente_id = c.id
                GROUP BY c.id, c.nombre
                ORDER BY ingresos DESC
                LIMIT 10
            ''')
            clientes_data = cursor.fetchall()
        
        resultado = []
        for row in clientes_data:
            promedio = row[2] / row[1] if row[1] > 0 else 0
            resultado.append({
                'nombre': row[0],
                'pedidos': row[1],
                'ingresos': round(float(row[2]), 2),
                'promedio': round(float(promedio), 2),
                'ultimo_pedido': row[3]
            })
        
        conn.close()
        print(f"👥 Mejores clientes encontrados: {len(resultado)}")
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Error en clientes-top: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/exportar', methods=['GET'])
def exportar_reporte():
    """Endpoint para exportar reportes a Excel"""
    periodo = request.args.get('periodo', 'mes')
    formato = request.args.get('formato', 'excel')
    
    if formato != 'excel':
        return jsonify({'error': 'Solo se soporta formato Excel'}), 400
    
    try:
        # Crear workbook
        wb = Workbook()
        
        # Obtener datos para todas las hojas
        inicio, fin, _, _ = get_periodo_fechas(periodo)
        conn = get_db_connection()
        
        # Hoja 1: Resumen general
        ws1 = wb.active
        ws1.title = "Resumen General"
        
        # Headers con estilo
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        ws1['A1'] = 'Métrica'
        ws1['B1'] = 'Valor'
        ws1['A1'].font = header_font
        ws1['B1'].font = header_font
        ws1['A1'].fill = header_fill
        ws1['B1'].fill = header_fill
        
        # Obtener datos del dashboard - PRIMERO INTENTAR CON VENTAS
        dashboard_data = conn.execute('''
            SELECT 
                COALESCE(SUM(total), 0) as ventas_totales,
                COUNT(*) as total_pedidos
            FROM ventas 
            WHERE fecha >= ? AND fecha <= ?
        ''', (inicio, fin)).fetchone()
        
        # FALLBACK: Si no hay ventas, usar pedidos
        if dashboard_data['total_pedidos'] == 0:
            dashboard_data = conn.execute('''
                SELECT 
                    COALESCE(SUM(pp.assigned_payment), 0) as ventas_totales,
                    COUNT(DISTINCT p.id) as total_pedidos
                FROM pedidos p
                LEFT JOIN pedido_productos pp ON p.id = pp.pedido_id
                WHERE p.fecha >= ? AND p.fecha <= ?
            ''', (inicio, fin)).fetchone()
            
        valor_promedio = dashboard_data['ventas_totales'] / dashboard_data['total_pedidos'] if dashboard_data['total_pedidos'] > 0 else 0
        
        # Clientes únicos - PRIMERO VENTAS, LUEGO PEDIDOS
        nuevos_clientes_query = conn.execute('''
            SELECT COUNT(DISTINCT cliente_id) as nuevos
            FROM ventas 
            WHERE fecha >= ? AND fecha <= ?
        ''', (inicio, fin)).fetchone()
        
        nuevos_clientes = nuevos_clientes_query['nuevos']
        if nuevos_clientes == 0:
            nuevos_clientes_query = conn.execute('''
                SELECT COUNT(DISTINCT cliente_id) as nuevos
                FROM pedidos 
                WHERE fecha >= ? AND fecha <= ?
            ''', (inicio, fin)).fetchone()
            nuevos_clientes = nuevos_clientes_query['nuevos']
        
        ws1['A2'] = 'Ventas Totales'
        ws1['B2'] = f"${dashboard_data['ventas_totales']:.2f}"
        ws1['A3'] = 'Total Pedidos'
        ws1['B3'] = dashboard_data['total_pedidos']
        ws1['A4'] = 'Valor Promedio'
        ws1['B4'] = f"${valor_promedio:.2f}"
        ws1['A5'] = 'Nuevos Clientes'
        ws1['B5'] = nuevos_clientes
        
        # Hoja 2: Ingresos por tipo
        ws2 = wb.create_sheet("Ingresos por Tipo")
        ws2['A1'] = 'Tipo'
        ws2['B1'] = 'Ingresos'
        ws2['C1'] = 'Porcentaje'
        ws2['A1'].font = header_font
        ws2['B1'].font = header_font
        ws2['C1'].font = header_font
        
        # Ingresos por tipo - PRIMERO VENTAS, LUEGO PEDIDOS
        ingresos_tipo = conn.execute('''
            SELECT 
                p.tipo,
                COALESCE(SUM(v.total), 0) as total_ingresos
            FROM ventas v
            JOIN productos p ON v.producto_id = p.id
            WHERE v.fecha >= ? AND v.fecha <= ?
            GROUP BY p.tipo
        ''', (inicio, fin)).fetchall()
        
        # FALLBACK: Si no hay datos en ventas, usar pedidos
        if not ingresos_tipo:
            ingresos_tipo = conn.execute('''
                SELECT 
                    pr.tipo,
                    COALESCE(SUM(pp.assigned_payment), 0) as total_ingresos
                FROM pedidos p
                JOIN pedido_productos pp ON p.id = pp.pedido_id
                JOIN productos pr ON pp.producto_id = pr.id
                WHERE p.fecha >= ? AND p.fecha <= ?
                GROUP BY pr.tipo
            ''', (inicio, fin)).fetchall()
        
        total_general = sum(row['total_ingresos'] for row in ingresos_tipo)
        row_num = 2
        for row in ingresos_tipo:
            porcentaje = (row['total_ingresos'] / total_general * 100) if total_general > 0 else 0
            ws2[f'A{row_num}'] = row['tipo'].upper()
            ws2[f'B{row_num}'] = f"${row['total_ingresos']:.2f}"
            ws2[f'C{row_num}'] = f"{porcentaje:.1f}%"
            row_num += 1
        
        # Hoja 3: Productos más vendidos
        ws3 = wb.create_sheet("Productos Top")
        ws3['A1'] = 'Producto'
        ws3['B1'] = 'Tipo'
        ws3['C1'] = 'Pedidos'
        ws3['D1'] = 'Ingresos'
        ws3['E1'] = 'Promedio'
        for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws3[col].font = header_font
        
        # Productos top - PRIMERO VENTAS, LUEGO PEDIDOS
        productos_top = conn.execute('''
            SELECT 
                p.nombre,
                p.tipo,
                COUNT(v.id) as pedidos,
                COALESCE(SUM(v.total), 0) as ingresos
            FROM ventas v
            JOIN productos p ON v.producto_id = p.id
            WHERE v.fecha >= ? AND v.fecha <= ?
            GROUP BY p.id, p.nombre, p.tipo
            ORDER BY ingresos DESC
            LIMIT 10
        ''', (inicio, fin)).fetchall()
        
        # FALLBACK: Si no hay ventas, usar pedidos
        if not productos_top:
            productos_top = conn.execute('''
                SELECT 
                    pr.nombre,
                    pr.tipo,
                    COUNT(p.id) as pedidos,
                    COALESCE(SUM(pp.assigned_payment), 0) as ingresos
                FROM pedidos p
                JOIN pedido_productos pp ON p.id = pp.pedido_id
                JOIN productos pr ON pp.producto_id = pr.id
                WHERE p.fecha >= ? AND p.fecha <= ?
                GROUP BY pr.id, pr.nombre, pr.tipo
                ORDER BY ingresos DESC
                LIMIT 10
            ''', (inicio, fin)).fetchall()
        
        row_num = 2
        for row in productos_top:
            promedio = row['ingresos'] / row['pedidos'] if row['pedidos'] > 0 else 0
            ws3[f'A{row_num}'] = row['nombre']
            ws3[f'B{row_num}'] = row['tipo'].upper()
            ws3[f'C{row_num}'] = row['pedidos']
            ws3[f'D{row_num}'] = f"${row['ingresos']:.2f}"
            ws3[f'E{row_num}'] = f"${promedio:.2f}"
            row_num += 1
        
        # Hoja 4: Mejores clientes
        ws4 = wb.create_sheet("Mejores Clientes")
        ws4['A1'] = 'Cliente'
        ws4['B1'] = 'Pedidos'
        ws4['C1'] = 'Ingresos'
        ws4['D1'] = 'Promedio'
        ws4['E1'] = 'Último Pedido'
        for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws4[col].font = header_font
        
        # Mejores clientes - PRIMERO VENTAS, LUEGO PEDIDOS
        clientes_top = conn.execute('''
            SELECT 
                c.nombre,
                COUNT(v.id) as pedidos,
                COALESCE(SUM(v.total), 0) as ingresos,
                MAX(v.fecha) as ultimo_pedido
            FROM ventas v
            JOIN clientes c ON v.cliente_id = c.id
            WHERE v.fecha >= ? AND v.fecha <= ?
            GROUP BY c.id, c.nombre
            ORDER BY ingresos DESC
            LIMIT 10
        ''', (inicio, fin)).fetchall()
        
        # FALLBACK: Si no hay ventas, usar pedidos
        if not clientes_top:
            clientes_top = conn.execute('''
                SELECT 
                    c.nombre,
                    COUNT(p.id) as pedidos,
                    COALESCE(SUM(pp.assigned_payment), 0) as ingresos,
                    MAX(p.fecha) as ultimo_pedido
                FROM pedidos p
                LEFT JOIN pedido_productos pp ON p.id = pp.pedido_id
                JOIN clientes c ON p.cliente_id = c.id
                WHERE p.fecha >= ? AND p.fecha <= ?
                GROUP BY c.id, c.nombre
                ORDER BY ingresos DESC
                LIMIT 10
            ''', (inicio, fin)).fetchall()
        
        row_num = 2
        for row in clientes_top:
            promedio = row['ingresos'] / row['pedidos'] if row['pedidos'] > 0 else 0
            ws4[f'A{row_num}'] = row['nombre']
            ws4[f'B{row_num}'] = row['pedidos']
            ws4[f'C{row_num}'] = f"${row['ingresos']:.2f}"
            ws4[f'D{row_num}'] = f"${promedio:.2f}"
            ws4[f'E{row_num}'] = row['ultimo_pedido']
            row_num += 1
        
        conn.close()
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"reporte_plusgraphics_{periodo}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/')
def landing():
    """Landing page mientras se carga frontend"""
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Plus Graphics - Sistema de Gestión</title>
        <style>
            body { font-family: Arial, sans-serif; max-width: 800px; margin: 50px auto; padding: 20px; }
            .btn { background: #007bff; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; display: inline-block; margin: 10px; }
            .btn:hover { background: #0056b3; }
            h1 { color: #333; }
            .credentials { background: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0; }
        </style>
    </head>
    <body>
        <h1>🎯 Plus Graphics - Sistema de Gestión</h1>
        <p>Sistema integral de gestión empresarial para servicios GFX/VFX</p>
        
        <div class="credentials">
            <h3>📋 Credenciales de Acceso:</h3>
            <p><strong>Administrador:</strong><br>
            📧 admin@plusgraphics.com<br>
            🔐 PlusGraphics2024!</p>
            
            <p><strong>Empleado:</strong><br>
            📧 empleado@plusgraphics.com<br>
            🔐 Empleado2024!</p>
        </div>
        
        <a href="/api/test" class="btn">🔧 Probar API</a>
        <a href="/api/reportes/dashboard?periodo=mes" class="btn">📊 Ver Estadísticas</a>
        
        <h3>📱 Módulos Disponibles via API:</h3>
        <ul>
            <li>👥 Gestión de Clientes</li>
            <li>📦 Catálogo de Productos/Servicios</li>
            <li>🛒 Control de Pedidos</li>
            <li>💰 Registro de Ventas</li>
            <li>📈 Cuentas por Cobrar</li>
            <li>📉 Cuentas por Pagar</li>
            <li>📊 Reportes y Analytics</li>
        </ul>
        
        <p><small>Sistema desarrollado para Plus Graphics - Empresa GFX/VFX</small></p>
    </body>
    </html>
    '''

@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    """Estadisticas dashboard - DATOS REALES"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Verificar si las tablas existen primero
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [row[0] for row in cursor.fetchall()]
        print(f"📋 Tablas encontradas: {tables}")
        
        # 1. Ganancias totales (suma de TODAS las ventas)
        ganancias_totales = 0
        if 'ventas' in tables:
            cursor.execute('SELECT COALESCE(SUM(total), 0) FROM ventas')
            ganancias_totales = cursor.fetchone()[0]
        
        # 2. Entregas pendientes (pedidos NO completados)
        entregas_pendientes = 0
        if 'pedidos' in tables:
            cursor.execute('''
                SELECT COUNT(*) FROM pedidos 
                WHERE estado NOT IN ("completado", "entregado", "finalizado")
            ''')
            entregas_pendientes = cursor.fetchone()[0] or 0
        
        # 3. Servicios disponibles (productos) - SIN FILTRO ESTADO
        servicios_disponibles = 0
        if 'productos' in tables:
            cursor.execute('SELECT COUNT(*) FROM productos')
            servicios_disponibles = cursor.fetchone()[0] or 0
            print(f"📦 Productos encontrados: {servicios_disponibles}")
        
        # 4. Total por pagar (solo si tabla existe)
        total_por_pagar = 0
        if 'cuentas_por_pagar' in tables:
            try:
                cursor.execute('''
                    SELECT COALESCE(SUM(saldo), 0) FROM cuentas_por_pagar 
                    WHERE estado = "pendiente"
                ''')
                total_por_pagar = cursor.fetchone()[0] or 0
            except Exception as e:
                print(f"⚠️ Error en cuentas_por_pagar: {str(e)}")
                total_por_pagar = 0
        
        # 5. Total por cobrar (solo si tabla existe)
        total_por_cobrar = 0
        if 'cuentas_por_cobrar' in tables:
            try:
                cursor.execute('''
                    SELECT COALESCE(SUM(saldo), 0) FROM cuentas_por_cobrar 
                    WHERE estado = "pendiente"
                ''')
                total_por_cobrar = cursor.fetchone()[0] or 0
            except Exception as e:
                print(f"⚠️ Error en cuentas_por_cobrar: {str(e)}")
                total_por_cobrar = 0
        
        # 6. Facturas vencidas (solo si tabla existe)
        facturas_vencidas = 0
        if 'cuentas_por_pagar' in tables:
            try:
                cursor.execute('''
                    SELECT COUNT(*) FROM cuentas_por_pagar 
                    WHERE estado = "pendiente" AND fecha_vencimiento < date('now')
                ''')
                facturas_vencidas = cursor.fetchone()[0] or 0
            except Exception as e:
                print(f"⚠️ Error calculando facturas vencidas: {str(e)}")
                facturas_vencidas = 0
        
        # 7. Pedidos recientes (solo si tablas existen)
        pedidos_recientes = []
        if 'pedidos' in tables and 'clientes' in tables and 'productos' in tables:
            try:
                cursor.execute('''
                    SELECT p.id, c.nombre, pr.nombre, p.total, p.fecha, p.estado
                    FROM pedidos p 
                    LEFT JOIN clientes c ON p.cliente_id = c.id
                    LEFT JOIN productos pr ON p.producto_id = pr.id
                    ORDER BY p.fecha DESC LIMIT 5
                ''')
                pedidos_data = cursor.fetchall()
                pedidos_recientes = [
                    {
                        'id': row[0],
                        'cliente': row[1] or 'Sin nombre',
                        'producto': row[2] or 'Sin producto',
                        'total': float(row[3]) if row[3] else 0,
                        'fecha': row[4],
                        'estado': row[5] or 'pendiente'
                    } for row in pedidos_data
                ]
            except Exception as e:
                print(f"⚠️ Error en pedidos recientes: {str(e)}")
                pedidos_recientes = []
        
        conn.close()
        
        result = {
            'ganancias_totales': float(ganancias_totales),
            'entregas_pendientes': entregas_pendientes,
            'servicios_disponibles': servicios_disponibles,
            'total_por_pagar': float(total_por_pagar),
            'total_por_cobrar': float(total_por_cobrar),
            'facturas_vencidas': facturas_vencidas,
            'pedidos_recientes': pedidos_recientes
        }
        
        print(f"📊 Dashboard stats calculadas: {result}")
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Error en dashboard stats: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/debug/database', methods=['GET'])
def debug_database():
    """Debug: verificar contenido base de datos"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Contar registros en cada tabla
        tables = ['clientes', 'productos', 'pedidos', 'ventas', 'cuentas_por_cobrar', 'cuentas_por_pagar']
        counts = {}
        
        for table in tables:
            cursor.execute(f'SELECT COUNT(*) FROM {table}')
            counts[table] = cursor.fetchone()[0]
        
        # Verificar datos sample
        cursor.execute('SELECT * FROM ventas LIMIT 3')
        ventas_sample = cursor.fetchall()
        
        cursor.execute('SELECT * FROM pedidos LIMIT 3')
        pedidos_sample = cursor.fetchall()
        
        conn.close()
        
        return jsonify({
            'tabla_counts': counts,
            'ventas_sample': [dict(row) for row in ventas_sample],
            'pedidos_sample': [dict(row) for row in pedidos_sample],
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/system/diagnosis', methods=['GET'])
def system_diagnosis():
    """Diagnostico completo del sistema"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # Verificar todas las tablas y contenido
        diagnosis = {
            'timestamp': datetime.now().isoformat(),
            'database_status': 'connected'
        }
        
        # 1. Verificar estructura tablas
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [row[0] for row in cursor.fetchall()]
        diagnosis['tables_found'] = tables
        
        # 2. Contar registros en cada tabla
        table_counts = {}
        for table in tables:
            cursor.execute(f'SELECT COUNT(*) FROM {table}')
            table_counts[table] = cursor.fetchone()[0]
        diagnosis['table_counts'] = table_counts
        
        # 3. Verificar datos específicos
        
        # CLIENTES
        cursor.execute('SELECT id, nombre FROM clientes ORDER BY id DESC LIMIT 5')
        clientes_data = cursor.fetchall()
        diagnosis['clientes_sample'] = [{'id': r[0], 'nombre': r[1]} for r in clientes_data]
        
        # PRODUCTOS
        cursor.execute('SELECT id, nombre, precio, estado FROM productos LIMIT 5')
        productos_data = cursor.fetchall()
        diagnosis['productos_sample'] = [{'id': r[0], 'nombre': r[1], 'precio': r[2], 'estado': r[3]} for r in productos_data]
        
        # PEDIDOS
        cursor.execute('SELECT id, cliente_id, producto_id, total, estado, fecha FROM pedidos ORDER BY id DESC LIMIT 5')
        pedidos_data = cursor.fetchall()
        diagnosis['pedidos_sample'] = [{'id': r[0], 'cliente_id': r[1], 'producto_id': r[2], 'total': r[3], 'estado': r[4], 'fecha': r[5]} for r in pedidos_data]
        
        # VENTAS
        cursor.execute('SELECT id, cliente_id, producto_id, total, fecha FROM ventas ORDER BY id DESC LIMIT 5')
        ventas_data = cursor.fetchall()
        diagnosis['ventas_sample'] = [{'id': r[0], 'cliente_id': r[1], 'producto_id': r[2], 'total': r[3], 'fecha': r[4]} for r in ventas_data]
        
        # 4. Calcular estadísticas manualmente
        
        # Total ventas
        cursor.execute('SELECT SUM(total) FROM ventas')
        total_ventas = cursor.fetchone()[0] or 0
        
        # Total productos activos
        cursor.execute('SELECT COUNT(*) FROM productos WHERE estado = "activo" OR estado = "Activo"')
        productos_activos = cursor.fetchone()[0] or 0
        
        # Pedidos pendientes
        cursor.execute('SELECT COUNT(*) FROM pedidos WHERE estado != "completado" AND estado != "entregado"')
        pedidos_pendientes = cursor.fetchone()[0] or 0
        
        # Cuentas por pagar
        cursor.execute('SELECT SUM(saldo) FROM cuentas_por_pagar WHERE estado = "pendiente"')
        total_por_pagar = cursor.fetchone()[0] or 0
        
        # Cuentas por cobrar
        cursor.execute('SELECT SUM(saldo) FROM cuentas_por_cobrar WHERE estado = "pendiente"')
        total_por_cobrar = cursor.fetchone()[0] or 0
        
        diagnosis['calculated_stats'] = {
            'total_ventas': total_ventas,
            'productos_activos': productos_activos,
            'pedidos_pendientes': pedidos_pendientes,
            'total_por_pagar': total_por_pagar,
            'total_por_cobrar': total_por_cobrar
        }
        
        conn.close()
        
        return jsonify(diagnosis)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'database_status': 'error',
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/api/test')
def api_test():
    """Endpoint de prueba para verificar funcionamiento"""
    return jsonify({
        'status': 'success',
        'message': 'API Plus Graphics funcionando correctamente',
        'modules': ['clientes', 'productos', 'pedidos', 'ventas', 'cuentas-cobrar', 'cuentas-pagar', 'reportes'],
        'timestamp': datetime.now().isoformat()
    })

if __name__ == '__main__':
    # Verificar seguridad antes de iniciar
    if not verificar_variables_entorno():
        exit(1)
    
    init_db()
    
    # Configuracion simple para Render
    import os
    port = int(os.environ.get('PORT', 5000))
    host = '0.0.0.0'
    debug = False  # Siempre False en produccion
    
    print(f"🚀 Plus Graphics Backend iniciando en puerto {port}")
    print(f"🌐 CORS configurado para Vercel")
    print(f"📊 Base de datos inicializada")
    
    app.run(host=host, port=port, debug=debug)

